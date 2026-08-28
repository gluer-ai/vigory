"""One-time seed: parse milinteltaxonomyontology.xlsx and upsert ClassDef,
LinkDef, VocabValue, ConfidenceCode nodes (+ uniqueness constraints) into Neo4j.

Usage: python ontology/import_ontology.py [path/to/xlsx]
Reads NEO4J_URI / NEO4J_USER / NEO4J_PASSWORD from .env at repo root.
"""
import sys
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

REPO_ROOT = Path(__file__).resolve().parent.parent
load_dotenv(REPO_ROOT / ".env")

import os

from neo4j import GraphDatabase

DEFAULT_XLSX = Path(__file__).resolve().parent / "milinteltaxonomyontology.xlsx"

CONSTRAINTS = [
    "CREATE CONSTRAINT classdef_key IF NOT EXISTS FOR (c:ClassDef) REQUIRE c.key IS UNIQUE",
    "CREATE CONSTRAINT linkdef_type IF NOT EXISTS FOR (l:LinkDef) REQUIRE l.type IS UNIQUE",
    "CREATE CONSTRAINT entity_id IF NOT EXISTS FOR (e:Entity) REQUIRE e.entity_id IS UNIQUE",
    "CREATE CONSTRAINT link_id IF NOT EXISTS FOR (l:Link) REQUIRE l.link_id IS UNIQUE",
]


def rows(ws):
    """Yield header-keyed dict rows, skipping the header itself."""
    it = ws.iter_rows(values_only=True)
    header = next(it)
    for r in it:
        if all(v is None for v in r):
            continue
        yield dict(zip(header, r))


def load_workbook(path: Path):
    return openpyxl.load_workbook(path, data_only=True)


def class_defs(wb):
    out = []
    for row in rows(wb["Entity_Taxonomy"]):
        full_key = row["Full_Key"]
        depth = row["Depth"]
        levels = full_key.split(".")
        parent_key = ".".join(levels[:-1]) if len(levels) > 1 else None
        label = levels[-1].replace("_", " ").title()
        out.append(
            {
                "key": full_key,
                "parent_key": parent_key,
                "level": depth,
                "label": label,
                "notes": row.get("Examples_Notes"),
            }
        )
    return out


def link_defs(wb):
    out = []
    for row in rows(wb["Link_Ontology"]):
        out.append(
            {
                "type": row["Link_Type"],
                "category": row.get("Category"),
                "domain": row["Domain (source class)"],
                "range": row["Range (target class)"],
                "directionality": row.get("Directionality"),
                "inverse": row.get("Inverse_Link"),
                "symmetric": row.get("Symmetric"),
                "transitive": row.get("Transitive"),
                "notes": row.get("Notes"),
            }
        )
    return out


def vocab_values(wb):
    out = []
    for row in rows(wb["Controlled_Vocabularies"]):
        out.append(
            {
                "list_name": row["List_Name"],
                "value": row["Value"],
                "applies_to": row.get("Applies_To"),
                "notes": row.get("Notes"),
            }
        )
    return out


def confidence_codes(wb):
    out = []
    for row in rows(wb["Confidence_Codes"]):
        out.append(
            {
                "code": row["Code"],
                "scale": row["Scale"],
                "meaning": row["Meaning"],
                "definition": row.get("Definition"),
            }
        )
    return out


def run_import(driver, wb):
    with driver.session() as session:
        for stmt in CONSTRAINTS:
            session.run(stmt)

        session.run(
            """
            UNWIND $rows AS row
            MERGE (c:ClassDef {key: row.key})
            SET c += row
            """,
            rows=class_defs(wb),
        )
        # single-inheritance parent edges, once all ClassDef nodes exist
        session.run(
            """
            MATCH (c:ClassDef) WHERE c.parent_key IS NOT NULL
            MATCH (p:ClassDef {key: c.parent_key})
            MERGE (c)-[:SUBCLASS_OF]->(p)
            """
        )

        session.run(
            """
            UNWIND $rows AS row
            MERGE (l:LinkDef {type: row.type})
            SET l += row
            """,
            rows=link_defs(wb),
        )

        session.run(
            """
            UNWIND $rows AS row
            MERGE (v:VocabValue {list_name: row.list_name, value: row.value})
            SET v += row
            """,
            rows=vocab_values(wb),
        )

        session.run(
            """
            UNWIND $rows AS row
            MERGE (cc:ConfidenceCode {code: row.code, scale: row.scale})
            SET cc += row
            """,
            rows=confidence_codes(wb),
        )

        counts = session.run(
            """
            RETURN
              count { MATCH (c:ClassDef) RETURN c } AS classes,
              count { MATCH (l:LinkDef) RETURN l } AS links,
              count { MATCH (v:VocabValue) RETURN v } AS vocab,
              count { MATCH (cc:ConfidenceCode) RETURN cc } AS confidence
            """
        ).single()
        return dict(counts)


def main():
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    wb = load_workbook(xlsx_path)

    uri = os.environ.get("NEO4J_URI", "bolt://localhost:7687")
    user = os.environ.get("NEO4J_USER", "neo4j")
    password = os.environ.get("NEO4J_PASSWORD", "changeme123")

    driver = GraphDatabase.driver(uri, auth=(user, password))
    try:
        counts = run_import(driver, wb)
        print(f"Seeded ontology into {uri}: {counts}")
    finally:
        driver.close()


if __name__ == "__main__":
    main()
